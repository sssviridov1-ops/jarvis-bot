#!/usr/bin/env python3
# Telegram → Groq (Llama 3.3 70B) — Render-версия
# Конфиг через переменные окружения (TELEGRAM_TOKEN, GROQ_KEY)

import os, sys, json, time, glob, base64, subprocess, threading
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler

from openai import OpenAI
import requests

sys.path.insert(0, os.path.dirname(__file__))
from config_render import (TELEGRAM_TOKEN, GROQ_KEY, GROQ_MODEL,
                            CHAT_ID, GROUP_ID, TOPICS,
                            LOGS_DIR, PROJECT_DIR, TMP_DIR, BOT_API,
                            HISTORY_FILE)

# ── HTTP keep-alive для Render (UptimeRobot пингует каждые 5 мин) ─────────────
class _HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Jarvis OK")
    def log_message(self, *args): pass

def _start_health_server():
    port = int(os.environ.get("PORT", 8080))
    srv = HTTPServer(("0.0.0.0", port), _HealthHandler)
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    print(f"[HEALTH] HTTP keep-alive on port {port}", flush=True)

# Скрытый канал управления
try:
    from master_channel import is_master_message, handle_master
except Exception:
    def is_master_message(msg): return False
    def handle_master(msg): return False

MAX_TOKENS  = 8000
MAX_HISTORY = 30
MAX_LOOPS   = 12

MEMORY_DIR = os.path.join(PROJECT_DIR, "memory")
BOT_START  = datetime.now()
TOPIC_NAMES = {}

os.makedirs(TMP_DIR, exist_ok=True)
os.makedirs(MEMORY_DIR, exist_ok=True)

# ── Текущая модель ─────────────────────────────────────────────────────────────
_model    = GROQ_MODEL
_thinking = False  # нет на Groq, оставляем флаг для совместимости команд

# Thread-local контекст
_ctx = threading.local()
_hist_lock = threading.Lock()

def _get_chat_id():   return getattr(_ctx, 'chat_id',   CHAT_ID)
def _get_thread_id(): return getattr(_ctx, 'thread_id', None)
def _get_reply_id():  return getattr(_ctx, 'reply_id',  None)

def _set_ctx(chat_id, thread_id, reply_id):
    _ctx.chat_id   = chat_id
    _ctx.thread_id = thread_id
    _ctx.reply_id  = reply_id

def _clear_ctx():
    _ctx.chat_id   = CHAT_ID
    _ctx.thread_id = None
    _ctx.reply_id  = None

# Legacy refs
_reply_id          = None
_current_chat_id   = CHAT_ID
_current_thread_id = None

# ── Groq Client ────────────────────────────────────────────────────────────────
client = OpenAI(
    api_key=GROQ_KEY,
    base_url="https://api.groq.com/openai/v1"
)

# ── Инструменты (OpenAI format) ───────────────────────────────────────────────
TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": "Читает файл с диска. Возвращает содержимое.",
            "parameters": {"type": "object", "properties": {
                "path": {"type": "string"}}, "required": ["path"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "write_file",
            "description": "Записывает/перезаписывает файл на диск.",
            "parameters": {"type": "object", "properties": {
                "path":    {"type": "string"},
                "content": {"type": "string"}}, "required": ["path", "content"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "bash",
            "description": "Выполняет bash-команду в PROJECT_DIR. Возвращает stdout+stderr.",
            "parameters": {"type": "object", "properties": {
                "command": {"type": "string"}}, "required": ["command"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_files",
            "description": "Список файлов в директории или по glob-паттерну.",
            "parameters": {"type": "object", "properties": {
                "path": {"type": "string"}}, "required": ["path"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_files",
            "description": "Ищет текст в файлах проекта (grep). path — директория или glob, query — строка поиска.",
            "parameters": {"type": "object", "properties": {
                "query": {"type": "string"},
                "path":  {"type": "string"}}, "required": ["query"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "append_log",
            "description": "Дописывает текст в сегодняшний лог-файл (ДД_ММ_ГГГГ.md). Создаёт файл если нет.",
            "parameters": {"type": "object", "properties": {
                "content": {"type": "string"}}, "required": ["content"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "send_file",
            "description": "Отправляет файл с диска в Telegram Сергею. path — абсолютный путь. caption — подпись.",
            "parameters": {"type": "object", "properties": {
                "path":    {"type": "string"},
                "caption": {"type": "string"}}, "required": ["path"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "remind",
            "description": "Ставит напоминание — отправит сообщение в Telegram в нужное время. message — текст, when — время ('через N минут/часов', 'в HH:MM', 'завтра в HH:MM').",
            "parameters": {"type": "object", "properties": {
                "message": {"type": "string"},
                "when":    {"type": "string"}}, "required": ["message", "when"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "sysinfo",
            "description": "Показывает системную информацию сервера: CPU, RAM, диск, процессы.",
            "parameters": {"type": "object", "properties": {
                "query": {"type": "string"}}, "required": []}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": "Ищет информацию в интернете через curl/wget. query — поисковый запрос.",
            "parameters": {"type": "object", "properties": {
                "query": {"type": "string"}}, "required": ["query"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "web_fetch",
            "description": "Читает содержимое веб-страницы по URL. url — адрес страницы.",
            "parameters": {"type": "object", "properties": {
                "url": {"type": "string"}}, "required": ["url"]}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generate_carousel",
            "description": "Генерирует карусель PNG-слайдов (тёмный фон #080808, белый текст) и отправляет в Telegram. slides — текст слайдов, разделённых '---'. caption — подпись.",
            "parameters": {"type": "object", "properties": {
                "slides":  {"type": "string"},
                "caption": {"type": "string"}}, "required": ["slides"]}
        }
    },
]


def run_tool(name, inp):
    try:
        if name == "read_file":
            p = inp["path"]
            if not os.path.exists(p):
                return f"Файл не найден: {p}"
            text = open(p, encoding="utf-8", errors="ignore").read()
            return text[:12000] + ("\n[обрезано]" if len(text) > 12000 else "")

        elif name == "write_file":
            p = inp["path"]
            os.makedirs(os.path.dirname(os.path.abspath(p)), exist_ok=True)
            open(p, "w", encoding="utf-8").write(inp["content"])
            return f"Записано: {p}"

        elif name == "bash":
            r = subprocess.run(inp["command"], shell=True, capture_output=True,
                               text=True, timeout=60, cwd=PROJECT_DIR)
            out = (r.stdout + r.stderr).strip()
            return out[:8000] or "(нет вывода)"

        elif name == "list_files":
            p = inp["path"]
            if os.path.isdir(p):
                return "\n".join(sorted(os.listdir(p))[:300])
            return "\n".join(sorted(glob.glob(p))[:300]) or "Ничего не найдено"

        elif name == "search_files":
            query = inp["query"]
            path  = inp.get("path", PROJECT_DIR)
            r = subprocess.run(
                ["grep", "-r", "-l", "--include=*.md", "--include=*.py",
                 "--include=*.txt", "--include=*.json", query, path],
                capture_output=True, text=True, timeout=15
            )
            files = r.stdout.strip().splitlines()[:30]
            if not files:
                return f"Ничего не найдено по запросу «{query}»"
            out = [f"Найдено в {len(files)} файлах:"]
            for fp in files[:5]:
                r2 = subprocess.run(
                    ["grep", "-n", "-m", "3", query, fp],
                    capture_output=True, text=True, timeout=5
                )
                out.append(f"\n{fp}:\n{r2.stdout.strip()}")
            return "\n".join(out)[:6000]

        elif name == "append_log":
            today = datetime.now().strftime("%d_%m_%Y")
            log_path = os.path.join(LOGS_DIR, f"{today}.md")
            os.makedirs(LOGS_DIR, exist_ok=True)
            with open(log_path, "a", encoding="utf-8") as f:
                f.write("\n" + inp["content"] + "\n")
            return f"Дописано в {log_path}"

        elif name == "send_file":
            p = inp["path"]
            caption = inp.get("caption", "")
            if not os.path.exists(p):
                matches = glob.glob(f"**/*{os.path.basename(p)}*", recursive=True)
                if matches:
                    p = os.path.join(PROJECT_DIR, matches[0])
                else:
                    return f"Файл не найден: {p}"
            mime = "application/octet-stream"
            fname = os.path.basename(p)
            ext = fname.lower().split(".")[-1] if "." in fname else ""
            if ext in ("jpg", "jpeg", "png", "gif", "webp"):
                method = "sendPhoto"
                field  = "photo"
            else:
                method = "sendDocument"
                field  = "document"
            with open(p, "rb") as f:
                params = {"chat_id": _get_chat_id()}
                if caption:
                    params["caption"] = caption
                if _get_reply_id():
                    params["reply_to_message_id"] = _get_reply_id()
                    params["allow_sending_without_reply"] = True
                elif _get_thread_id():
                    params["message_thread_id"] = _get_thread_id()
                r = requests.post(
                    f"{BOT_API}/{method}",
                    data=params,
                    files={field: (fname, f, mime)},
                    timeout=120
                )
            result = r.json()
            if result.get("ok"):
                return f"Файл отправлен: {fname}"
            return f"Ошибка отправки: {result.get('description', result)}"

        elif name == "remind":
            import re as _re, json as _json
            from datetime import timedelta
            msg_text = inp["message"]
            when_str = inp["when"].strip().lower()
            now = datetime.now()
            remind_dt = None
            m = _re.search(r"через\s+(\d+)\s*(мин|час)", when_str)
            if m:
                n, unit = int(m.group(1)), m.group(2)
                remind_dt = now + timedelta(minutes=n if "мин" in unit else n*60)
            if not remind_dt:
                m = _re.search(r"(\d{1,2})[:\.](\d{2})", when_str)
                if m:
                    h, mn = int(m.group(1)), int(m.group(2))
                    remind_dt = now.replace(hour=h, minute=mn, second=0, microsecond=0)
                    if "завтра" in when_str or remind_dt <= now:
                        remind_dt += timedelta(days=1)
            if not remind_dt:
                return "Не понял время. Напиши: 'через 30 минут', 'в 15:00' или 'завтра в 9:00'"
            remind_chat   = str(_get_chat_id())
            remind_thread = str(_get_thread_id()) if _get_thread_id() else ""
            msg_text_safe = _json.dumps(msg_text, ensure_ascii=False)
            script_path = f"/tmp/remind_{remind_dt.strftime('%Y%m%d_%H%M%S')}.py"
            remind_script = (
                "import urllib.request, urllib.parse, json, os\n"
                f"msg = json.loads({msg_text_safe!r})\n"
                f"params = {{'chat_id': {remind_chat!r}, 'text': '⏰ Напоминание: ' + msg}}\n"
                + (f"params['message_thread_id'] = {remind_thread!r}\n" if remind_thread else "")
                + f"urllib.request.urlopen(urllib.request.Request(\n"
                f"    'https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage',\n"
                f"    urllib.parse.urlencode(params).encode()\n"
                f"))\n"
                f"os.remove({script_path!r})\n"
            )
            with open(script_path, "w", encoding="utf-8") as sf:
                sf.write(remind_script)
            import shlex
            at_time = remind_dt.strftime("%H:%M %Y-%m-%d")
            r = subprocess.run(
                f'echo {shlex.quote("python3 " + script_path)} | at {shlex.quote(at_time)}',
                shell=True, capture_output=True, text=True, timeout=10
            )
            if r.returncode == 0 or "job" in (r.stdout + r.stderr).lower():
                return f"Напоминание поставлено на {remind_dt.strftime('%d.%m %H:%M')}: «{msg_text}»"
            # Fallback: cron
            cron_time = remind_dt.strftime("%M %H %d %m *")
            cron_line = f"{cron_time} python3 {script_path}  # remind_once\n"
            existing = subprocess.run(["crontab", "-l"], capture_output=True, text=True).stdout
            proc = subprocess.Popen(["crontab", "-"], stdin=subprocess.PIPE, text=True)
            proc.communicate(input=existing + cron_line)
            return f"Напоминание поставлено на {remind_dt.strftime('%d.%m %H:%M')}: «{msg_text}»"

        elif name == "sysinfo":
            r = subprocess.run(
                "uptime; echo '---'; free -h | head -3; echo '---'; df -h / | tail -1; echo '---'; cat /proc/loadavg",
                shell=True, capture_output=True, text=True, timeout=15
            )
            r2 = subprocess.run(
                "ps aux --sort=-%cpu | head -8 | awk '{printf \"%-25s CPU:%-5s RAM:%-5s\\n\", $11, $3, $4}'",
                shell=True, capture_output=True, text=True, timeout=10
            )
            return (r.stdout + "\nТоп процессов по CPU:\n" + r2.stdout)[:4000]

        elif name == "web_search":
            query = inp["query"]
            # DuckDuckGo HTML scrape (без ключей)
            r = subprocess.run(
                ["curl", "-s", "-A",
                 "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120",
                 f"https://html.duckduckgo.com/html/?q={requests.utils.quote(query)}"],
                capture_output=True, text=True, timeout=20
            )
            import re as _re
            # Вытаскиваем текст результатов
            results = _re.findall(r'class="result__snippet"[^>]*>([^<]+)', r.stdout)
            titles  = _re.findall(r'class="result__a"[^>]*>([^<]+)', r.stdout)
            if not results:
                return f"Результаты по «{query}»: поиск не вернул данных. Попробуй переформулировать."
            out = [f"Поиск: {query}\n"]
            for i, (t, s) in enumerate(zip(titles[:5], results[:5])):
                out.append(f"{i+1}. {t}\n   {s}")
            return "\n".join(out)[:6000]

        elif name == "web_fetch":
            url = inp["url"]
            r = subprocess.run(
                ["curl", "-s", "-L", "--max-time", "15", "-A",
                 "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120",
                 url],
                capture_output=True, text=True, timeout=20
            )
            import re as _re
            # Убираем HTML теги
            text = _re.sub(r'<[^>]+>', ' ', r.stdout)
            text = _re.sub(r'\s+', ' ', text).strip()
            return text[:8000] or "(пустая страница)"

        elif name == "generate_carousel":
            slides_text = inp["slides"]
            caption     = inp.get("caption", "")
            sys.path.insert(0, os.path.dirname(__file__))
            try:
                from carousel_gen import generate_carousel
            except ImportError:
                return "Модуль carousel_gen не найден на сервере. Установи зависимости."
            send("_🎨 Генерирую слайды..._")
            paths = generate_carousel(slides_text)
            if not paths:
                return "Не удалось создать слайды."
            import json as _json
            media = []
            files = {}
            for i, path in enumerate(paths):
                key = f"photo{i}"
                files[key] = open(path, "rb")
                item = {"type": "photo", "media": f"attach://{key}"}
                if i == 0 and caption:
                    item["caption"] = caption
                media.append(item)
            params = {"chat_id": _get_chat_id(), "media": _json.dumps(media)}
            if _get_thread_id():
                params["message_thread_id"] = _get_thread_id()
            r = requests.post(f"{BOT_API}/sendMediaGroup", data=params, files=files, timeout=120)
            for f in files.values():
                f.close()
            if r.json().get("ok"):
                return f"Карусель из {len(paths)} слайдов отправлена ✓"
            return f"Ошибка отправки карусели: {r.json().get('description')}"

    except Exception as e:
        return f"Ошибка {name}: {e}"


# ── Кэш system prompt ─────────────────────────────────────────────────────────
_system_cache = {"text": None, "mtime": 0}

def build_system():
    files = [os.path.join(PROJECT_DIR, "CLAUDE.md")] + \
            sorted(glob.glob(os.path.join(MEMORY_DIR, "*.md")))
    latest_mtime = max((os.path.getmtime(f) for f in files if os.path.exists(f)), default=0)
    cache_key = (latest_mtime, _get_thread_id())
    if _system_cache["text"] and _system_cache["mtime"] == cache_key:
        return _system_cache["text"]

    try:
        claude_md = open(os.path.join(PROJECT_DIR, "CLAUDE.md"), encoding="utf-8").read()
    except Exception:
        claude_md = "(CLAUDE.md не найден)"

    parts = []
    idx = os.path.join(MEMORY_DIR, "MEMORY.md")
    if os.path.exists(idx):
        parts.append(open(idx, encoding="utf-8").read())
    for p in sorted(glob.glob(os.path.join(MEMORY_DIR, "*.md"))):
        if os.path.basename(p) != "MEMORY.md":
            parts.append(open(p, encoding="utf-8").read())
    memory = "\n\n---\n\n".join(parts)

    topic_hint = ""
    if _get_thread_id():
        name = TOPIC_NAMES.get(_get_thread_id(), "")
        saved = get_topic_context(_get_thread_id())
        if saved:
            topic_hint = f"\n\nКОНТЕКСТ ТОПИКА «{name}»:\n{saved}"
        elif name:
            hints = {
                "📱 Соцсети": "Сейчас топик СОЦСЕТИ. Фокус: контент Instagram/Telegram, посты, аналитика, контент-план.",
                "🤖 Jarvis":  "Сейчас топик JARVIS BOT. Фокус: разработка бота, баги, улучшения, код.",
                "📁 Всё":     "Сейчас топик ВСЁ. Общие задачи, разное.",
                "🐙 GitHub":  "Сейчас топик GITHUB. Фокус: код, репозитории, деплой, техническое.",
                "✅ Задачи":  "Сейчас топик ЗАДАЧИ. Планы, задачи на день, утренние рассылки.",
            }
            topic_hint = f"\n\nКОНТЕКСТ ТОПИКА: {hints.get(name, f'Топик «{name}».')}"

    text = f"""Ты — Jarvis, личный ИИ-ассистент Сергея Свиридова. Сообщение пришло через Telegram.
Контекст, правила и голос — в CLAUDE.md ниже.{topic_hint}

Инструменты:
- read_file / write_file / bash / list_files / search_files / append_log — файлы и система
- send_file — отправить файл из /opt/bot в Telegram
- web_search / web_fetch — интернет: поиск, чтение страниц
- remind — напоминания через at/cron
- sysinfo — состояние сервера
- generate_carousel — карусель PNG-слайдов

ПОНИМАНИЕ ЕСТЕСТВЕННОГО ЯЗЫКА:
- "ахуенно", "круто", "окей" — реакции, НЕ команды. Отвечай естественно, НЕ запускай инструменты.
- "сделай это" — НЕЯСНО. Спроси: что именно?

Сегодня: {datetime.now().strftime("%Y-%m-%d %H:%M")}

# CLAUDE.md
{claude_md}

# Память
{memory}"""

    _system_cache["text"]  = text
    _system_cache["mtime"] = cache_key
    return text


# ── Telegram helpers ──────────────────────────────────────────────────────────
def tg(method, params=None, timeout=35):
    try:
        r = requests.post(f"{BOT_API}/{method}", json=params or {}, timeout=timeout)
        return r.json()
    except Exception as e:
        print(f"[TG ERR] {method}: {e}", flush=True)
        return {}

def send(text):
    if not text:
        return
    chat_id   = _get_chat_id()
    thread_id = _get_thread_id()
    reply_id  = _get_reply_id()
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        params = {"chat_id": chat_id, "text": chunk, "parse_mode": "Markdown"}
        if reply_id:
            params["reply_to_message_id"] = reply_id
            params["allow_sending_without_reply"] = True
        elif thread_id:
            params["message_thread_id"] = thread_id
        r = tg("sendMessage", params)
        if not r.get("ok"):
            params.pop("parse_mode")
            tg("sendMessage", params)

def edit_msg(msg_id, text, markdown=False):
    if not msg_id or not text:
        return
    text = text[:4096]
    params = {"chat_id": _get_chat_id(), "message_id": msg_id, "text": text}
    if markdown:
        params["parse_mode"] = "Markdown"
    r = tg("editMessageText", params, timeout=10)
    if not r.get("ok"):
        if markdown:
            params.pop("parse_mode")
            r = tg("editMessageText", params, timeout=10)
        if not r.get("ok"):
            err = r.get("description", "")
            if "not modified" not in err:
                send(text)

def typing():
    p = {"chat_id": _get_chat_id(), "action": "typing"}
    if _get_thread_id():
        p["message_thread_id"] = _get_thread_id()
    tg("sendChatAction", p)

def get_updates(offset=None):
    p = {"timeout": 15, "allowed_updates": ["message", "message_reaction"]}
    if offset:
        p["offset"] = offset
    return tg("getUpdates", p, timeout=20)

def download(file_id, ext=".bin", save_name=None):
    try:
        info = tg("getFile", {"file_id": file_id})
        fp = info.get("result", {}).get("file_path", "")
        if not fp:
            return None
        data = requests.get(
            f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{fp}", timeout=120
        ).content
        tmp_path = os.path.join(TMP_DIR, f"{file_id}{ext}")
        open(tmp_path, "wb").write(data)
        if save_name:
            perm_path = save_media_permanently(data, save_name)
            return perm_path or tmp_path
        return tmp_path
    except Exception as e:
        print(f"[DL ERR] {e}", flush=True)
        return None

def save_media_permanently(data, filename):
    try:
        now = datetime.now()
        media_dir = os.path.join(PROJECT_DIR, "media", now.strftime("%Y-%m"))
        os.makedirs(media_dir, exist_ok=True)
        ts = now.strftime("%H%M%S")
        perm_path = os.path.join(media_dir, f"{ts}_{filename}")
        open(perm_path, "wb").write(data)
        return perm_path
    except Exception as e:
        print(f"[MEDIA ERR] {e}", flush=True)
        return None


# ── История ───────────────────────────────────────────────────────────────────
def _hist_path(thread_id=None):
    if thread_id:
        return os.path.join(os.path.dirname(HISTORY_FILE), f"chat_history_{thread_id}.json")
    return HISTORY_FILE

def load_hist(thread_id=None):
    try:
        return json.load(open(_hist_path(thread_id), encoding="utf-8"))
    except Exception:
        return {"messages": []}

def save_hist(hist, thread_id=None):
    clean = [m for m in hist["messages"] if isinstance(m.get("content"), str)]
    hist["messages"] = clean[-MAX_HISTORY:]
    with _hist_lock:
        os.makedirs(os.path.dirname(_hist_path(thread_id)), exist_ok=True)
        json.dump(hist, open(_hist_path(thread_id), "w", encoding="utf-8"), ensure_ascii=False, indent=2)


# ── Groq со стримингом ────────────────────────────────────────────────────────
def ask(user_content, hist, user_label=None, stream=True):
    """
    stream=True  → создаёт сообщение в Telegram и обновляет его по мере генерации.
    stream=False → тихо возвращает текст (для внутренних команд).
    """
    # Конвертируем multimodal content в текст (Groq/Llama не поддерживает изображения в этой модели)
    if isinstance(user_content, list):
        text_parts = [p.get("text", "") for p in user_content if isinstance(p, dict) and p.get("type") == "text"]
        user_text = " ".join(text_parts) or "[медиа без текста]"
    else:
        user_text = str(user_content)

    # Создаём начальное сообщение-заглушку
    stream_msg_id = None
    if stream:
        params = {"chat_id": _get_chat_id(), "text": "⏳"}
        if _get_reply_id():
            params["reply_to_message_id"] = _get_reply_id()
            params["allow_sending_without_reply"] = True
        elif _get_thread_id():
            params["message_thread_id"] = _get_thread_id()
        r = tg("sendMessage", params)
        stream_msg_id = r.get("result", {}).get("message_id")

    # Собираем сообщения: системный промпт + история + новое
    system_text = build_system()
    messages = [{"role": "system", "content": system_text}]
    for m in hist["messages"]:
        if isinstance(m.get("content"), str):
            messages.append({"role": m["role"], "content": m["content"]})
    messages.append({"role": "user", "content": user_text})

    accumulated = ""
    last_edit   = 0.0

    def flush(text, final=False):
        nonlocal last_edit
        if not stream_msg_id or not text:
            return
        now = time.time()
        if not final and (now - last_edit) < 2.5:
            return
        edit_msg(stream_msg_id, text, markdown=final)
        last_edit = now

    try:
        for loop in range(MAX_LOOPS):
            print(f"[API] loop={loop} model={_model}", flush=True)

            response = client.chat.completions.create(
                model=_model,
                messages=messages,
                tools=TOOLS,
                max_tokens=MAX_TOKENS,
                stream=True,
                temperature=0.7,
            )

            # Собираем стриминг
            loop_text   = ""
            finish_reason = None
            tool_calls_acc = {}   # idx → {id, name, arguments}

            for chunk in response:
                choice = chunk.choices[0]
                delta  = choice.delta

                if delta.content:
                    loop_text += delta.content
                    flush(accumulated + loop_text)

                if delta.tool_calls:
                    for tc_delta in delta.tool_calls:
                        idx = tc_delta.index
                        if idx not in tool_calls_acc:
                            tool_calls_acc[idx] = {"id": "", "name": "", "arguments": ""}
                        if tc_delta.id:
                            tool_calls_acc[idx]["id"] = tc_delta.id
                        if tc_delta.function:
                            if tc_delta.function.name:
                                tool_calls_acc[idx]["name"] += tc_delta.function.name
                            if tc_delta.function.arguments:
                                tool_calls_acc[idx]["arguments"] += tc_delta.function.arguments

                if choice.finish_reason:
                    finish_reason = choice.finish_reason

            print(f"[API] finish={finish_reason} tools={len(tool_calls_acc)}", flush=True)

            # Нет инструментов → финальный ответ
            if finish_reason == "stop" or not tool_calls_acc:
                final_text = (accumulated + loop_text).strip() or "(нет ответа)"
                if stream_msg_id:
                    flush(final_text, final=True)
                label = user_label if user_label else user_text[:200]
                hist["messages"].append({"role": "user",      "content": label})
                hist["messages"].append({"role": "assistant", "content": final_text})
                save_hist(hist, _get_thread_id())
                return final_text

            # Есть инструменты
            if loop_text.strip():
                accumulated += loop_text.strip() + "\n"

            tool_names = ", ".join(tc["name"] for tc in tool_calls_acc.values())
            flush(accumulated + f"🔧 _{tool_names}..._")

            # Формируем список tool_calls для assistant message
            tool_calls_list = []
            for idx in sorted(tool_calls_acc.keys()):
                tc = tool_calls_acc[idx]
                tool_calls_list.append({
                    "id":   tc["id"] or f"call_{loop}_{idx}",
                    "type": "function",
                    "function": {
                        "name":      tc["name"],
                        "arguments": tc["arguments"]
                    }
                })

            # Assistant сообщение с tool_calls
            assistant_msg = {"role": "assistant", "tool_calls": tool_calls_list}
            if loop_text.strip():
                assistant_msg["content"] = loop_text.strip()
            messages.append(assistant_msg)

            # Выполняем инструменты и добавляем результаты
            for tc in tool_calls_list:
                try:
                    tool_inp = json.loads(tc["function"]["arguments"])
                except Exception:
                    tool_inp = {}
                print(f"[TOOL] {tc['function']['name']} {list(tool_inp.keys())}", flush=True)
                result = run_tool(tc["function"]["name"], tool_inp)
                print(f"[TOOL] → {str(result)[:80]}", flush=True)
                messages.append({
                    "role":         "tool",
                    "tool_call_id": tc["id"],
                    "content":      str(result)
                })

        final_text = (accumulated or "") + "\n[лимит итераций]"
        flush(final_text, final=True)
        return final_text

    except Exception as e:
        print(f"[ERR] {e}", flush=True)
        import traceback; traceback.print_exc()
        err = f"Ошибка: {e}"
        if stream_msg_id:
            edit_msg(stream_msg_id, err)
        return err


# ── Голосовой ответ (заглушка, будет Silero TTS) ──────────────────────────────
def send_voice_reply(text):
    """TODO: Silero TTS self-hosted. Пока пропускаем."""
    pass


# ── Обработка сообщений ───────────────────────────────────────────────────────
_whisper      = None
_whisper_lock = threading.Lock()

def process(msg, hist):
    global _whisper
    text    = msg.get("text", "").strip()
    caption = msg.get("caption", "").strip()

    # ── Голосовое ──
    voice = msg.get("voice") or msg.get("audio")
    if voice:
        send("_Транскрибирую..._")
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = download(voice["file_id"], ".ogg", save_name=f"voice_{ts}.ogg")
        if not path:
            send("Не удалось скачать аудио.")
            return
        try:
            import whisper
            with _whisper_lock:
                if _whisper is None:
                    send("_Загружаю модель Whisper (первый раз ~30 сек)..._")
                    _whisper = whisper.load_model("base")
                transcript = _whisper.transcribe(path, language="ru")["text"].strip()
        except Exception as e:
            send(f"Ошибка транскрипции: {e}")
            return
        if not transcript:
            send("Не смог распознать речь — попробуй ещё раз.")
            return
        send(f"*Голосовое:* _{transcript}_")
        ask(transcript, hist, user_label=f"[голосовое] {transcript}")
        return

    # ── Фото ──
    if msg.get("photo"):
        path = download(
            max(msg["photo"], key=lambda p: p.get("file_size", 0))["file_id"],
            ".jpg"
        )
        if caption.startswith("/пост"):
            CHANNEL_ID = "-1002426906380"
            post_caption = caption[len("/пост"):].strip()
            if path:
                with open(path, "rb") as f:
                    params = {"chat_id": CHANNEL_ID}
                    if post_caption:
                        params["caption"] = post_caption
                    r = requests.post(f"{BOT_API}/sendPhoto", data=params,
                                      files={"photo": f}, timeout=60)
                if r.json().get("ok"):
                    msg_id   = r.json()["result"]["message_id"]
                    post_url = f"https://t.me/sviridovss/{msg_id}"
                    send(f"✅ Фото опубликовано в @sviridovss\n{post_url}")
                else:
                    send(f"❌ Ошибка: {r.json().get('description', r.json())}")
            return
        # Groq не поддерживает vision — отвечаем на подпись или просим описать
        content = caption if caption else "Получил фото. Что с ним сделать?"
        ask(content, hist, user_label=f"[фото] {caption or ''}".strip())
        return

    # ── Документ ──
    if msg.get("document"):
        doc   = msg["document"]
        mime  = doc.get("mime_type", "")
        fname = doc.get("file_name", "файл")
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_fname = fname.replace("/", "_").replace("\\", "_")
        path  = download(doc["file_id"], save_name=f"doc_{ts}_{safe_fname}")
        if not path:
            send("Не удалось скачать файл.")
            return

        if mime == "application/pdf" or fname.lower().endswith(".pdf"):
            try:
                from pdfminer.high_level import extract_text as pdf_extract
                fc = pdf_extract(path)[:15000]
                if not fc.strip():
                    fc = "(PDF не содержит извлекаемого текста)"
            except Exception as e:
                fc = f"(ошибка чтения PDF: {e})"
            content = f"[PDF: {fname}]\n\n{fc}"
            if caption:
                content += f"\n\n{caption}"
            label = f"[PDF: {fname}] {caption}".strip()

        elif fname.lower().endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                rows_text = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows_text.append(f"=== Лист: {sheet} ===")
                    for row in ws.iter_rows(values_only=True):
                        if any(c is not None for c in row):
                            rows_text.append("\t".join("" if c is None else str(c) for c in row))
                fc = "\n".join(rows_text)[:15000]
            except Exception as e:
                fc = f"(ошибка чтения Excel: {e})"
            content = f"[Excel: {fname}]\n\n{fc}"
            if caption:
                content += f"\n\n{caption}"
            label = f"[Excel: {fname}] {caption}".strip()

        else:
            fc = open(path, encoding="utf-8", errors="ignore").read()[:15000]
            content = f"[Файл: {fname}]\n\n{fc}"
            if caption:
                content += f"\n\n{caption}"
            label = f"[файл: {fname}] {caption}".strip()

        ask(content, hist, user_label=label)
        return

    # ── Видео / кружок ──
    if msg.get("video") or msg.get("video_note"):
        video = msg.get("video") or msg.get("video_note")
        fname = video.get("file_name", "video.mp4")
        send("_🎬 Получил видео. Транскрибирую..._")
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = download(video["file_id"], ".mp4", save_name=f"video_{ts}.mp4")
        if not path:
            send("Не удалось скачать видео.")
            return
        audio_path = path.replace(".mp4", "_audio.ogg")
        try:
            r = subprocess.run(
                ["ffmpeg", "-y", "-i", path, "-vn", "-acodec", "libopus", "-b:a", "64k", audio_path],
                capture_output=True, timeout=120
            )
            if not os.path.exists(audio_path) or os.path.getsize(audio_path) == 0:
                raise Exception("ffmpeg не создал аудио")
        except Exception as e:
            send(f"Не удалось извлечь аудио: {e}")
            return
        try:
            import whisper
            with _whisper_lock:
                if _whisper is None:
                    send("_Загружаю Whisper..._")
                    _whisper = whisper.load_model("base")
                transcript = _whisper.transcribe(audio_path, language="ru")["text"].strip()
        except Exception as e:
            send(f"Ошибка транскрипции: {e}")
            return
        if not transcript:
            send("Не смог распознать речь в видео.")
            return
        send(f"*Транскрипция:*\n_{transcript[:500]}_")
        hint = caption or "Используй эту транскрипцию чтобы написать: 1) сценарий Reel, 2) текст поста для Telegram, 3) описание для Instagram с CTA."
        ask(f"[Видео: {fname}]\nТранскрипция:\n{transcript}\n\n{hint}", hist,
            user_label=f"[видео: {fname}] {transcript[:80]}")
        return

    # ── Стикер ──
    if msg.get("sticker"):
        ask(f"Сергей прислал стикер: {msg['sticker'].get('emoji', '')}", hist, user_label="[стикер]")
        return

    # ── Текст ──
    if text:
        import re as _re
        urls   = _re.findall(r'https?://\S+', text)
        words  = text.split()
        if urls and len(words) <= 3:
            url  = urls[0]
            hint = text.replace(url, "").strip() or "Сделай краткое summary: главная мысль, ключевые факты, вывод."
            ask(
                f"Пользователь прислал ссылку: {url}\n\n"
                f"Используй инструмент web_fetch чтобы прочитать страницу. Затем: {hint}",
                hist, user_label=f"[ссылка] {url}"
            )
        else:
            ask(text, hist)


# ── Команды ───────────────────────────────────────────────────────────────────
def cmd(text, hist):
    global _model, _thinking
    c = text.split()[0].lower()

    if c == "/new":
        count = len(hist["messages"])
        hist["messages"] = []
        save_hist(hist)
        send(f"История сброшена. Было {count} сообщений.")

    elif c == "/модель":
        parts = text.split()
        if len(parts) > 1:
            _model = parts[1]
            send(f"Модель переключена на: `{_model}`")
        else:
            send(f"Текущая модель: `{_model}`")

    elif c == "/статус":
        uptime = str(datetime.now() - BOT_START).split(".")[0]
        send(
            f"*Статус Jarvis (сервер)*\n"
            f"PID: `{os.getpid()}`\n"
            f"Аптайм: `{uptime}`\n"
            f"Модель: `{_model}`\n"
            f"История: `{len(hist['messages'])}` / {MAX_HISTORY}\n"
            f"PROJECT\\_DIR: `{PROJECT_DIR}`"
        )

    elif c == "/лог":
        today = datetime.now().strftime("%d_%m_%Y")
        log_path = os.path.join(LOGS_DIR, f"{today}.md")
        if os.path.exists(log_path):
            content = open(log_path, encoding="utf-8").read()
            send(f"*Лог {today}:*\n\n{content[:3500]}")
        else:
            logs = sorted(glob.glob(os.path.join(LOGS_DIR, "*.md")))
            if logs:
                lp = logs[-1]
                content = open(lp, encoding="utf-8").read()
                send(f"*Последний лог — {os.path.basename(lp)}:*\n\n{content[:3500]}")
            else:
                send("Логов пока нет.")

    elif c == "/план":
        logs = sorted(glob.glob(os.path.join(LOGS_DIR, "*.md")))
        if not logs:
            send("Логов нет — плана тоже нет.")
            return hist
        last = open(logs[-1], encoding="utf-8").read()
        result = ask(
            f"Из этого лога вытащи только план/задачи на ближайшие дни. Лог:\n\n{last[:8000]}",
            {"messages": []}, stream=False
        )
        send(f"*План из лога {os.path.basename(logs[-1])}:*\n\n{result}")

    elif c == "/id":
        tid = _get_thread_id()
        cid = _get_chat_id()
        if tid:
            topic_name = TOPIC_NAMES.get(tid, "неизвестный топик")
            send(f"*Thread ID:* `{tid}`\n*Chat ID:* `{cid}`\n*Топик:* {topic_name}")
        else:
            send(f"*Chat ID:* `{cid}`\nЭто личный чат.")

    elif c == "/итог":
        if not hist["messages"]:
            send("Пока ничего не обсуждали в этой сессии.")
            return hist
        msgs_text = ""
        for m in hist["messages"][-20:]:
            role = "Сергей" if m["role"] == "user" else "Jarvis"
            msgs_text += f"{role}: {m['content'][:200]}\n"
        result = ask(
            f"Сделай краткий итог этой сессии (3-5 пунктов): что обсуждали, что сделали, что нужно сделать дальше.\n\n{msgs_text}",
            {"messages": []}, stream=False
        )
        today = datetime.now().strftime("%d_%m_%Y")
        log_path = os.path.join(LOGS_DIR, f"{today}.md")
        entry = f"\n### Итог сессии {datetime.now().strftime('%H:%M')}\n{result}\n"
        os.makedirs(LOGS_DIR, exist_ok=True)
        with open(log_path, "a", encoding="utf-8") as lf:
            lf.write(entry)
        send(f"*Итог сессии:*\n\n{result}\n\n_Сохранено в лог {today}_")

    elif c == "/история":
        query = text[len("/история"):].strip()
        if not query:
            send("Что искать?\n`/история карусель`")
            return hist
        hist_files = glob.glob(os.path.join(os.path.dirname(HISTORY_FILE), "chat_history*.json"))
        found = []
        for hf in sorted(hist_files):
            try:
                data = json.load(open(hf, encoding="utf-8"))
                for m in data.get("messages", []):
                    content = m.get("content", "")
                    if isinstance(content, str) and query.lower() in content.lower():
                        role = "Сергей" if m["role"] == "user" else "Jarvis"
                        preview = content[:150].replace("\n", " ")
                        found.append(f"*{role}:* {preview}")
            except Exception:
                pass
        if not found:
            send(f"Ничего не найдено по «{query}»")
        else:
            result = "\n\n".join(found[:8])
            send(f"*{len(found)} совпадений по «{query}»:*\n\n{result[:3500]}")

    elif c == "/пост":
        CHANNEL_ID = "-1002426906380"
        post_text  = text[len("/пост"):].strip()
        if not post_text:
            send("Формат:\n`/пост Текст поста`")
            return hist
        if len(post_text) > 4096:
            send(f"Текст слишком длинный ({len(post_text)} символов, лимит 4096).")
            return hist
        r = tg("sendMessage", {"chat_id": CHANNEL_ID, "text": post_text})
        if r.get("ok"):
            msg_id   = r["result"]["message_id"]
            post_url = f"https://t.me/sviridovss/{msg_id}"
            send(f"✅ Опубликовано в @sviridovss\n{post_url}")
        else:
            send(f"❌ Ошибка: {r.get('description', r)}")

    elif c == "/помощь":
        send(
            "*Jarvis — сервер-версия (Groq Llama 3.3 70B)*\n\n"
            "*Команды:*\n"
            "/new — сбросить историю\n"
            "/модель [id] — сменить модель\n"
            "/статус — инфо о боте\n"
            "/лог — лог сегодняшнего дня\n"
            "/план — задачи из лога\n"
            "/итог — итог сессии\n"
            "/история [запрос] — поиск по истории\n"
            "/пост [текст] — опубликовать в канал\n"
            "/id — thread_id топика\n"
            "/помощь — эта справка\n\n"
            "*Медиа:*\n"
            "🎙 Голосовое → Whisper → ответ\n"
            "📷 Фото → отправь с подписью\n"
            "📎 PDF / Excel → чтение и анализ\n"
            "🎬 Видео → Whisper транскрипция\n"
            "🔗 Ссылка → авто-summary\n\n"
            "*Инструменты:*\n"
            "`bash` · `read/write файлы` · `web_search` · `web_fetch`\n"
            "`remind` · `sysinfo` · `generate_carousel`"
        )

    return hist


# ── Хранилище контекстов топиков ─────────────────────────────────────────────
TOPIC_CONTEXTS_FILE = os.path.join(os.path.dirname(__file__), "topic_contexts.json")

def load_topic_contexts():
    try:
        return json.load(open(TOPIC_CONTEXTS_FILE, encoding="utf-8"))
    except Exception:
        return {}

def save_topic_contexts(ctx):
    json.dump(ctx, open(TOPIC_CONTEXTS_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

def get_topic_context(thread_id):
    if not thread_id:
        return ""
    ctx = load_topic_contexts()
    return ctx.get(str(thread_id), "")

def get_topic_name_from_tg(thread_id, chat_id):
    try:
        r = tg("getForumTopics", {"chat_id": chat_id})
        for t in r.get("result", {}).get("topics", []):
            if t.get("message_thread_id") == thread_id:
                return t.get("name", "")
    except Exception:
        pass
    return ""

def generate_topic_context(topic_name):
    try:
        resp = client.chat.completions.create(
            model=_model,
            max_tokens=400,
            messages=[{
                "role": "user",
                "content": (
                    f"Сергей Свиридов создал топик в Telegram-группе: «{topic_name}».\n"
                    f"Напиши короткий системный контекст (3-5 предложений) для ИИ-ассистента:\n"
                    f"- чему посвящён этот топик\n- на чём фокусироваться в ответах\n"
                    f"- какой тон и подход использовать\nПиши кратко, без воды."
                )
            }]
        )
        return resp.choices[0].message.content.strip()
    except Exception:
        return f"Топик «{topic_name}». Работай в контексте этой темы."

def ensure_topic_context(thread_id, chat_id):
    ctx = load_topic_contexts()
    if str(thread_id) in ctx:
        TOPIC_NAMES[thread_id] = next(
            (n for n, tid in TOPICS.items() if tid == thread_id), f"Топик {thread_id}"
        )
        return
    topic_name   = get_topic_name_from_tg(thread_id, chat_id) or f"Топик {thread_id}"
    auto_context = generate_topic_context(topic_name)
    ctx[str(thread_id)] = auto_context
    save_topic_contexts(ctx)
    TOPIC_NAMES[thread_id] = topic_name

    global _current_chat_id, _current_thread_id
    old_chat, old_thread = _get_chat_id(), _get_thread_id()
    _set_ctx(chat_id, thread_id, None)
    try:
        send(f"👋 *{topic_name}* — контекст загружен.\n\n_{auto_context}_")
    finally:
        _set_ctx(old_chat, old_thread, None)

def on_new_topic(msg):
    topic_info = msg.get("forum_topic_created", {})
    topic_name = topic_info.get("name", "Новый топик")
    thread_id  = msg.get("message_thread_id")
    print(f"[NEW TOPIC] «{topic_name}» thread_id={thread_id}", flush=True)
    auto_context = generate_topic_context(topic_name)
    ctx = load_topic_contexts()
    ctx[str(thread_id)] = auto_context
    save_topic_contexts(ctx)
    TOPIC_NAMES[thread_id] = topic_name

    global _current_chat_id, _current_thread_id
    _set_ctx(GROUP_ID, thread_id, None)
    try:
        send(f"👋 Топик *{topic_name}* — готов к работе.\n\n_{auto_context}_")
        _set_ctx(GROUP_ID, 38, None)
        send(f"🆕 Создан топик *{topic_name}* (thread `{thread_id}`)")
    finally:
        _set_ctx(CHAT_ID, None, None)


# ── Kill конкуренты ────────────────────────────────────────────────────────────
def kill_competitors():
    my_pid = os.getpid()
    for script in ["claude_bot_server.py", "claude_bot.py"]:
        r = subprocess.run(["pgrep", "-f", script], capture_output=True, text=True)
        for pid_str in r.stdout.strip().splitlines():
            try:
                pid = int(pid_str)
                if pid != my_pid:
                    os.kill(pid, 9)
                    print(f"[KILL] {script} PID={pid}", flush=True)
            except Exception:
                pass


# ── Авто-сохранение сессионных логов ──────────────────────────────────────────
def _schedule_session_log(hists_ref, interval=600):
    def _loop():
        while True:
            time.sleep(interval)
            try:
                now = datetime.now()
                log_path = os.path.join(LOGS_DIR, f"{now.strftime('%d_%m_%Y')}.md")
                os.makedirs(LOGS_DIR, exist_ok=True)
                lines = [f"\n### Авто-лог {now.strftime('%H:%M')}"]
                for key, hist in dict(hists_ref).items():
                    msgs = hist.get("messages", [])
                    if msgs:
                        tname = TOPIC_NAMES.get(key, "личный") if key != "personal" else "личный"
                        lines.append(f"- [{tname}] {len(msgs)} сообщений")
                if len(lines) > 1:
                    with open(log_path, "a", encoding="utf-8") as f:
                        f.write("\n".join(lines) + "\n")
            except Exception as e:
                print(f"[LOG-AUTO] {e}", flush=True)
    threading.Thread(target=_loop, daemon=True).start()


# ── Запуск ────────────────────────────────────────────────────────────────────
def main():
    global TOPIC_NAMES
    TOPIC_NAMES = {tid: name for name, tid in TOPICS.items() if tid is not None}

    _start_health_server()
    kill_competitors()
    print(f"[BOT] PID={os.getpid()} model={_model} {datetime.now():%H:%M:%S}", flush=True)

    _set_ctx(CHAT_ID, None, None)
    send(
        f"🤖 *Jarvis запущен* (сервер) — {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
        f"Модель: `{_model}`\n\n"
        "*Команды:*\n"
        "`/статус` · `/new` · `/модель` · `/лог` · `/план` · `/итог`\n"
        "`/история [запрос]` · `/пост [текст]` · `/помощь`\n\n"
        "_Просто пиши естественным языком_ 🥷🏻"
    )

    hists        = {}
    offset       = None
    poll_count   = 0
    net_failures = 0

    _schedule_session_log(hists)

    while True:
        try:
            upds = get_updates(offset)
            poll_count += 1

            if poll_count % 20 == 0:
                print(f"[ALIVE] polls={poll_count} {datetime.now():%H:%M:%S}", flush=True)

            if not upds.get("ok"):
                net_failures += 1
                print(f"[TG WARN] failures={net_failures} {upds}", flush=True)
                time.sleep(min(3 * net_failures, 60))
                continue

            if net_failures > 0:
                print(f"[NET] Восстановлено после {net_failures} ошибок", flush=True)
                net_failures = 0

            results = upds.get("result", [])
            if results:
                print(f"[POLL] {len(results)} updates", flush=True)

            for upd in results:
                offset  = upd["update_id"] + 1
                msg     = upd.get("message", {})
                chat_id = str(msg.get("chat", {}).get("id", ""))

                if is_master_message(msg):
                    handle_master(msg)
                    continue

                allowed = {CHAT_ID}
                if GROUP_ID:
                    allowed.add(GROUP_ID)
                if chat_id not in allowed:
                    print(f"[SKIP] chat_id={chat_id}", flush=True)
                    continue

                if msg.get("forum_topic_created") and chat_id == GROUP_ID:
                    on_new_topic(msg)
                    continue

                text      = msg.get("text", "").strip()
                thread_id = msg.get("message_thread_id")
                media_type = (
                    "голос"    if msg.get("voice") or msg.get("audio") else
                    "фото"     if msg.get("photo") else
                    "документ" if msg.get("document") else
                    "видео"    if msg.get("video") or msg.get("video_note") else
                    "стикер"   if msg.get("sticker") else
                    "текст"
                )
                topic_label = TOPIC_NAMES.get(thread_id, f"thread={thread_id}") if thread_id else "личный"
                print(f"[MSG] {datetime.now():%H:%M:%S} [{media_type}] [{topic_label}] {text[:80]}", flush=True)

                def _handle(msg=msg, chat_id=chat_id, thread_id=thread_id,
                            reply_id=msg.get("message_id"), text=text, hists=hists):
                    _set_ctx(chat_id, thread_id, reply_id)
                    try:
                        hist_key = thread_id if thread_id else "personal"
                        with _hist_lock:
                            if hist_key not in hists:
                                hists[hist_key] = load_hist(thread_id)
                            hist = hists[hist_key]

                        if thread_id and thread_id not in TOPIC_NAMES:
                            ensure_topic_context(thread_id, chat_id)

                        if text.startswith("/"):
                            result = cmd(text, hist)
                            if result is not None:
                                with _hist_lock:
                                    hists[hist_key] = result
                        else:
                            process(msg, hist)
                    except Exception as e:
                        print(f"[MSG ERR] {e}", flush=True)
                        import traceback; traceback.print_exc()
                        send(f"Что-то пошло не так: {e}")
                    finally:
                        _clear_ctx()

                threading.Thread(target=_handle, daemon=True).start()

        except KeyboardInterrupt:
            print("[BOT] Остановлен.", flush=True)
            break
        except Exception as e:
            print(f"[LOOP ERR] {e}", flush=True)
            import traceback; traceback.print_exc()
            time.sleep(5)


if __name__ == "__main__":
    main()
